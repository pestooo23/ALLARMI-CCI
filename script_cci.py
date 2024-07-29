#--------------------------------------------------------------------------
#                           IMPORT LIBRRERIE

import os
import re
import pandas as pd
import extract_msg  
import logging  
import tkinter as tk  
from tkinter import filedialog  
from datetime import datetime  
from openpyxl import load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo



#------------------------------------------------------------------------
#                          LOG CONFIGURATION

def config_log(dest_folder):
    
    log_file = os.path.join(dest_folder, 'estrazione_allarmi.log')
    logging.basicConfig(
        filename=log_file,
        level=logging.DEBUG,  
        format='%(asctime)s - %(levelname)s - %(message)s'
    )
    return log_file



#------------------------------------------------------------------------
#               EXTRACT DATA FROM MESSAGE

def estrai_dati_da_messaggio(msg):
    
    body = msg.body  
    date = msg.date 
    logging.debug(f"Corpo del messaggio:\n{body}")
    logging.debug(f"Data del messaggio: {date}")

    
    date_naive = date.replace(tzinfo=None)
    
    date_formatted = date_naive.strftime("%d/%m/%Y %H:%M")

    
    luogo_pattern = re.compile(r'Allarme attivo\s+(.+?)\s+->')
    error_type_pattern = re.compile(r'->\s+(.+?)\s+Valore soglia')
    max_campioni_pattern = re.compile(r'Numero max campioni consecutivi raggiunti:\s*(\d+)')

    
    luoghi = luogo_pattern.findall(body)
    error_types = error_type_pattern.findall(body)
    max_campioni = max_campioni_pattern.findall(body)

    logging.debug(f"Luoghi trovati: {luoghi}")
    logging.debug(f"Error Types trovati: {error_types}")
    logging.debug(f"Max Campioni trovati: {max_campioni}")

    dati = []
    
    for luogo, error_type, max_campione in zip(luoghi, error_types, max_campioni):
        dati.append({
            'LUOGO': luogo.strip(),
            'ERROR TYPE': error_type.strip(),
            'MAX CAMPIONI': int(max_campione.strip()),
            'DATA': date_formatted  
        })
        logging.debug(f"Dati estratti: {dati[-1]}")

    return dati



#------------------------------------------------------------------------
#               ADJUST COLUMN WIDTHS

def adjust_column_widths(ws):
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter  # Get the column name
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width


#------------------------------------------------------------------------
#               RENAME SHEET WITH CURRENT DATE

def rinomina_foglio_con_data_corrente(ws):
    current_date = datetime.now().strftime("%d-%m")
    ws.title = current_date
    logging.info(f"Foglio rinominato in: {ws.title}")


#------------------------------------------------------------------------
#               PROCESSING FOLDERS

def elabora_cartelle(cartelle, dest_folder):
    dati_completi = []

    for cartella in cartelle:
        for file_name in os.listdir(cartella):
            if file_name.endswith('.msg'):
                file_path = os.path.join(cartella, file_name)
                msg = extract_msg.Message(file_path)
                dati_messaggio = estrai_dati_da_messaggio(msg)
                dati_completi.extend(dati_messaggio)

    if not dati_completi:
        logging.error("Nessun dato estratto. Verificare le espressioni regolari e il contenuto dei messaggi.")
    else:
        df = pd.DataFrame(dati_completi)

        output_file = os.path.join(dest_folder, 'output.xlsx')
        df.to_excel(output_file, index=False)

        wb = load_workbook(output_file)
        ws = wb.active

        min_col = ws.min_column
        min_row = ws.min_row
        max_col = ws.max_column
        max_row = ws.max_row

        tab = Table(displayName="DatiEstratti", ref=f"{ws.cell(min_row, min_col).coordinate}:{ws.cell(max_row, max_col).coordinate}")
        style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                               showLastColumn=False, showRowStripes=True, showColumnStripes=True)
        tab.tableStyleInfo = style
        ws.add_table(tab)

        adjust_column_widths(ws)  # Adjust column widths

        rinomina_foglio_con_data_corrente(ws)  # Rename sheet with current date

        wb.save(output_file)

        logging.info(f'Dati estratti e salvati in {output_file}')



#------------------------------------------------------------------------
#               SELECTING FOLDERS

def seleziona_cartelle():
    cartelle = []
    while True:
        cartella = filedialog.askdirectory()  
        if cartella:
            cartelle.append(cartella)  
            add_more = tk.messagebox.askyesno("Seleziona Cartelle", "Vuoi selezionare un'altra cartella?")
            if not add_more:
                break
        else:
            break
    if cartelle:
        path_entry.delete(0, tk.END)
        path_entry.insert(0, ','.join(cartelle)) 

def seleziona_destinazione():
    dest_folder = filedialog.askdirectory()  
    if dest_folder:
        dest_entry.delete(0, tk.END)
        dest_entry.insert(0, dest_folder)  



#------------------------------------------------------------------------
#               START EXTRACTION

def avvia_estrazione():
    cartelle = path_entry.get().split(',')
    dest_folder = dest_entry.get()  

    
    if all(os.path.isdir(cartella) for cartella in cartelle) and os.path.isdir(dest_folder):
        log_file = config_log(dest_folder)  
        elabora_cartelle(cartelle, dest_folder)  
    else:
        logging.error(f"Cartella non valida: {cartelle} o {dest_folder}")  



#------------------------------------------------------------------------
#               CREATING GUI 

root = tk.Tk()  
root.title("Estrazione Allarmi CCI")  

frame = tk.Frame(root)  
frame.pack(padx=10, pady=10)  


path_label = tk.Label(frame, text="Folder path:")
path_label.grid(row=0, column=0, padx=5, pady=5)

path_entry = tk.Entry(frame, width=50)
path_entry.grid(row=0, column=1, padx=5, pady=5)

browse_button = tk.Button(frame, text="Browse", command=seleziona_cartelle)
browse_button.grid(row=0, column=2, padx=5, pady=5)

dest_label = tk.Label(frame, text="Destination path:")
dest_label.grid(row=1, column=0, padx=5, pady=5)

dest_entry = tk.Entry(frame, width=50)
dest_entry.grid(row=1, column=1, padx=5, pady=5)

dest_button = tk.Button(frame, text="Browse", command=seleziona_destinazione)
dest_button.grid(row=1, column=2, padx=5, pady=5)

start_button = tk.Button(frame, text="Start extraction", command=avvia_estrazione)
start_button.grid(row=2, columnspan=3, pady=10)  

root.mainloop()
